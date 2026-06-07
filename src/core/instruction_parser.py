"""Parse the dialogue instruction Excel into structured InstructionSpec JSON.

Strategy:
- openpyxl reads rows; skip rows that are empty or contain only an id.
- For each valid markdown block, run a deterministic regex segmentation
  to recover section bodies even when the headings vary between Chinese
  and English templates.
- Then call an LLM to convert ``Call Flow`` / ``Conversation Flow`` into a
  flow graph and to mine constraints + FAQ key points. The LLM is asked
  for a strict JSON object that we validate against ``InstructionSpec``.
- Variables in the opening line template (e.g. ``${rider_name}``) are
  collected so the orchestrator can inject concrete values later.
"""
from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from .llm_client import ChatMessage, LLMClient, build_default_client
from .schemas import (
    FlowEdge,
    FlowNode,
    HardConstraints,
    InstructionConstraints,
    InstructionSpec,
    KnowledgePoint,
)

logger = logging.getLogger(__name__)


SECTION_HEADERS = {
    "role": [r"role"],
    "task": [r"task"],
    "opening_line": [r"opening\s*line"],
    "call_flow": [r"call\s*flow", r"conversation\s*flow", r"对话流程", r"会话流程"],
    "knowledge": [r"knowledge\s*points?\s*(\(faq\))?", r"faq", r"知识点"],
    "constraints": [r"constraints?", r"约束"],
}

VARIABLE_PATTERN = re.compile(r"\$\{([a-zA-Z_][a-zA-Z0-9_]*)\}")


def _split_sections(markdown: str) -> dict[str, str]:
    """Split markdown text into the canonical sections using header heuristics."""
    lines = markdown.splitlines()
    section_indices: list[tuple[str, int, str, str]] = []
    for i, line in enumerate(lines):
        m = re.match(r"^\s*(#{1,3})\s*([^\n]+?)\s*$", line)
        if not m:
            continue
        depth = len(m.group(1))
        body = m.group(2).strip()
        inline_value = ""
        header_part = body
        if ":" in body or "：" in body:
            sep_match = re.search(r"[:：]", body)
            if sep_match:
                header_part = body[: sep_match.start()].strip()
                inline_value = body[sep_match.end():].strip()
        header_clean = re.sub(r"^[0-9]+\.\s*", "", header_part.lower())
        matched: Optional[str] = None
        for canonical, patterns in SECTION_HEADERS.items():
            if any(re.match(p, header_clean, flags=re.IGNORECASE) for p in patterns):
                matched = canonical
                break
        if not matched:
            continue
        if depth > 2 and matched not in {"call_flow", "knowledge"}:
            continue
        section_indices.append((matched, i, inline_value, header_part))
    sections: dict[str, str] = {}
    for idx, (canonical, line_idx, inline_value, _) in enumerate(section_indices):
        end_idx = (
            section_indices[idx + 1][1] if idx + 1 < len(section_indices) else len(lines)
        )
        body_lines = lines[line_idx + 1 : end_idx]
        body = "\n".join(body_lines).strip()
        if inline_value and not body:
            body = inline_value
        elif inline_value and body:
            body = f"{inline_value}\n{body}"
        if canonical not in sections:
            sections[canonical] = body
    return sections


def _collect_variables(markdown: str) -> list[str]:
    seen: list[str] = []
    for m in VARIABLE_PATTERN.finditer(markdown):
        name = m.group(1)
        if name not in seen:
            seen.append(name)
    return seen


def _load_rows(xlsx_path: str | Path) -> list[tuple[str, str]]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[tuple[str, str]] = []
    for r in range(2, ws.max_row + 1):
        id_cell = ws.cell(row=r, column=1).value
        body_cell = ws.cell(row=r, column=2).value
        if id_cell is None and body_cell is None:
            continue
        body_text = str(body_cell).strip() if body_cell is not None else ""
        if not body_text:
            logger.warning("Skip row %s: empty instruction body", id_cell)
            continue
        rows.append((str(id_cell).strip(), body_text))
    return rows


SYSTEM_PROMPT = """你是一名对话流程的解析器。我将给你一段中文外呼对话任务指令（Markdown），\
请你抽取出结构化的流程图、知识点和约束。\n\n请只输出一个 JSON 对象，不要任何额外解释。\n\n要求：\n\
1. flow_nodes：流程中的关键节点。每个节点 id 用 S1/S2/...，desc 简洁中文描述。\n\
2. flow_edges：节点之间的有向跳转，condition 用中文描述触发条件（例如"用户是负责人"、"用户不知情"）。\n\
3. knowledge：每条 FAQ 抽出 topic、触发词 triggers（中文关键词数组）、key_points（应当覆盖的要点中文数组）。\n\
4. constraints.hard 中收集硬性可机器判定项：\n\
   - max_chars_per_reply: 整数（例如指令说"每次回复约30字以内"则填30；若给出区间取上限）。\n\
   - forbidden_words: 不允许出现的词。\n\
   - no_discount_promise: 是否禁止承诺折扣/优惠券，bool。\n\
   - required_out_of_scope_reply: 若指令规定遇到越权问题的固定话术，写出该话术原文，否则 null。\n\
   - opening_keywords: 开场白模板中应当包含的关键短语数组。\n\
   - required_replies: 对话中触发条件下要求的固定话术数组，每项 {trigger, reply, when}。\n\
5. constraints.soft：需要 LLM 评判的软约束（语气随意、避免重复、自然过渡等）。\n\
6. constraints.termination：终止/挂断策略数组（什么情况下要求挂断或结束）。\n\
"""

USER_TEMPLATE = """以下是任务指令的 Markdown 原文，请输出 JSON：\n\n```markdown\n{markdown}\n```\n\n字段示例：\n\
{{\n  "flow_nodes": [{{"id":"S1","desc":"..."}}],\n  "flow_edges": [{{"source":"S1","target":"S2","condition":"..."}}],\n  "knowledge": [{{"topic":"...","triggers":["..."],"key_points":["..."]}}],\n  "constraints": {{"hard": {{...}}, "soft": ["..."], "termination": ["..."]}}\n}}\n"""


def _extract_json(text: str) -> dict[str, Any]:
    if not text:
        raise ValueError("Empty LLM response")
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?", "", cleaned).strip()
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3].strip()
    start = cleaned.find("{")
    end = cleaned.rfind("}")
    if start == -1 or end == -1 or end < start:
        raise ValueError(f"No JSON object in response: {text!r}")
    return json.loads(cleaned[start : end + 1])


def _extract_first_line(text: str) -> str:
    for line in text.splitlines():
        s = line.strip()
        if s:
            return s
    return ""


def _build_spec(
    instruction_id: str,
    markdown: str,
    sections: dict[str, str],
    llm_data: dict[str, Any],
) -> InstructionSpec:
    flow_nodes = [FlowNode(**n) for n in llm_data.get("flow_nodes", [])]
    flow_edges = [FlowEdge(**e) for e in llm_data.get("flow_edges", [])]
    knowledge = [KnowledgePoint(**k) for k in llm_data.get("knowledge", [])]
    constraints_data = llm_data.get("constraints", {}) or {}
    hard_data = constraints_data.get("hard", {}) or {}
    hard = HardConstraints(**hard_data)
    constraints = InstructionConstraints(
        hard=hard,
        soft=list(constraints_data.get("soft", []) or []),
        termination=list(constraints_data.get("termination", []) or []),
    )
    opening_line = sections.get("opening_line", "").strip()
    if not opening_line:
        opening_line = _extract_first_line(sections.get("call_flow", ""))
    role = sections.get("role", "").strip() or llm_data.get("role", "")
    task = sections.get("task", "").strip() or llm_data.get("task", "")
    return InstructionSpec(
        id=instruction_id,
        role=role,
        task=task,
        opening_line_template=opening_line,
        variables=_collect_variables(markdown),
        flow_nodes=flow_nodes,
        flow_edges=flow_edges,
        knowledge=knowledge,
        constraints=constraints,
        raw_markdown=markdown,
    )


CHAR_LIMIT_PATTERN = re.compile(r"(?:约|约为|不超过|最多|控制在)?\s*(\d{2,3})\s*(?:个)?字")


def _offline_flow_from_section(call_flow_text: str) -> tuple[list[FlowNode], list[FlowEdge]]:
    """Heuristic parsing of Call Flow sections without LLM assistance.

    Two heading styles are supported:

    1. Numbered top-level steps: ``1. ...`` at the beginning of a line.
    2. Markdown-style ``## Step N: ...`` headers (newer instructions).
    """
    if not call_flow_text:
        return [], []
    lines = call_flow_text.splitlines()
    has_explicit_steps = any(
        re.match(r"^\s*#{2,3}\s*Step\s*\d+", line, re.IGNORECASE) for line in lines
    )
    nodes: list[FlowNode] = []
    edges: list[FlowEdge] = []
    counter = 0
    last_id: Optional[str] = None
    if has_explicit_steps:
        pattern = re.compile(
            r"^\s*#{2,3}\s*Step\s*(\d+)\s*[:：]?\s*(.+?)\s*$", re.IGNORECASE
        )
        for raw_line in lines:
            m = pattern.match(raw_line)
            if not m:
                continue
            counter += 1
            node_id = f"S{counter}"
            desc = m.group(2).strip().rstrip(":：。").strip()
            nodes.append(FlowNode(id=node_id, desc=desc[:80]))
            if last_id is not None:
                edges.append(
                    FlowEdge(source=last_id, target=node_id, condition="顺序进入")
                )
            last_id = node_id
    else:
        pattern = re.compile(r"^(\d+)[\.、:：]\s*(.+)$")
        for raw_line in lines:
            line = raw_line.strip()
            m = pattern.match(line)
            if not m:
                continue
            counter += 1
            node_id = f"S{counter}"
            desc = m.group(2).strip().rstrip(":：。").strip()
            nodes.append(FlowNode(id=node_id, desc=desc[:80]))
            if last_id is not None:
                edges.append(
                    FlowEdge(source=last_id, target=node_id, condition="顺序进入")
                )
            last_id = node_id
    if not nodes:
        nodes.append(FlowNode(id="S1", desc="按指令完成对话"))
    return nodes, edges


def _offline_knowledge(knowledge_text: str) -> list[KnowledgePoint]:
    points: list[KnowledgePoint] = []
    for raw in re.split(r"^\s*[-*]\s+", knowledge_text, flags=re.MULTILINE):
        item = raw.strip()
        if not item:
            continue
        topic_candidate = item.split("：")[0].split(":")[0].strip()
        topic = re.sub(r"\*+", "", topic_candidate).strip()[:30]
        triggers = []
        for kw in re.findall(r"\*\*([^*]+)\*\*", item):
            kw = re.sub(r"\s+", "", kw).strip()
            if kw and kw not in triggers:
                triggers.append(kw)
        points.append(
            KnowledgePoint(
                topic=topic or "未命名要点",
                triggers=triggers,
                key_points=[item[:200]],
            )
        )
    return points


def _offline_constraints(constraints_text: str) -> InstructionConstraints:
    hard = HardConstraints()
    soft: list[str] = []
    termination: list[str] = []
    lines = [
        l.strip().lstrip("-*").strip()
        for l in constraints_text.splitlines()
        if l.strip()
    ]
    for line in lines:
        lower = line.lower()
        if "字" in line and CHAR_LIMIT_PATTERN.search(line):
            m = CHAR_LIMIT_PATTERN.search(line)
            if m:
                hard.max_chars_per_reply = int(m.group(1))
        if any(kw in line for kw in ("不能承诺", "不要承诺", "不承诺")) and any(
            kw in line for kw in ("折扣", "优惠券", "返利")
        ):
            hard.no_discount_promise = True
        if "禁止" in line and ":" in line:
            try:
                rhs = line.split(":", 1)[1]
                words = [w.strip(" ，,、；;。") for w in re.split(r"[、,，]", rhs) if w.strip()]
                hard.forbidden_words.extend(words)
            except Exception:
                pass
        if any(kw in line for kw in ("开车", "在开车", "我在开车")) and "挂断" in line:
            termination.append(line)
        if "坚持" in line and "挂断" in line:
            termination.append(line)
        if any(kw in line for kw in ("不说", "禁用")) and "：" in line:
            try:
                rhs = line.split("：", 1)[1]
                for token in re.findall(r"[\"'\u201c]([^\"'\u201d]+)[\"'\u201d]", rhs):
                    if token and token not in hard.forbidden_words:
                        hard.forbidden_words.append(token)
            except Exception:
                pass
        if any(kw in line for kw in ("职责", "无法回答", "超出")):
            quoted = re.findall(r"[\"\u201c]([^\"\u201d]+)[\"\u201d]", line)
            if quoted:
                hard.required_out_of_scope_reply = quoted[0]
        if any(kw in line for kw in ("语气", "自然", "随意", "重复", "礼貌", "过渡")):
            soft.append(line)
    return InstructionConstraints(hard=hard, soft=soft, termination=termination)


def _offline_opening_keywords(opening_line: str) -> list[str]:
    if not opening_line:
        return []
    text = opening_line.strip()
    parts = re.split(r"[。！？!?；;]", text)
    keywords: list[str] = []
    for part in parts:
        p = part.strip()
        if not p:
            continue
        if VARIABLE_PATTERN.search(p):
            for fragment in VARIABLE_PATTERN.split(p)[::2]:
                fragment = re.sub(r"\*+", "", fragment)
                fragment = re.sub(r"\s+", "", fragment).strip("，,。！？!?；;：:")
                if 4 <= len(fragment) <= 40 and fragment not in keywords:
                    keywords.append(fragment[:40])
            if len(keywords) >= 3:
                break
            continue
        cleaned = p
        cleaned = re.sub(r"\*+", "", cleaned)
        cleaned = re.sub(r"\s+", "", cleaned).strip()
        if 4 <= len(cleaned) <= 40:
            keywords.append(cleaned[:40])
        if len(keywords) >= 3:
            break
    return keywords


def parse_instruction_offline(instruction_id: str, markdown: str) -> InstructionSpec:
    sections = _split_sections(markdown)
    call_flow_text = sections.get("call_flow", "")
    flow_nodes, flow_edges = _offline_flow_from_section(call_flow_text)
    knowledge = _offline_knowledge(sections.get("knowledge", ""))
    if not knowledge and call_flow_text:
        knowledge = _knowledge_from_call_flow(call_flow_text)
    constraints = _offline_constraints(sections.get("constraints", ""))
    opening_line = sections.get("opening_line", "").strip()
    constraints.hard.opening_keywords = _offline_opening_keywords(opening_line)
    role = sections.get("role", "").strip()
    task = sections.get("task", "").strip()
    return InstructionSpec(
        id=instruction_id,
        role=role,
        task=task,
        opening_line_template=opening_line,
        variables=_collect_variables(markdown),
        flow_nodes=flow_nodes,
        flow_edges=flow_edges,
        knowledge=knowledge,
        constraints=constraints,
        raw_markdown=markdown,
    )


def _knowledge_from_call_flow(call_flow_text: str) -> list[KnowledgePoint]:
    """For instructions like #2 that embed FAQ-like info in the flow, mine 参考话术 and bullets."""
    points: list[KnowledgePoint] = []
    for m in re.finditer(
        r"\*{0,2}参考话术[：:]\*{0,2}\s*(.+)", call_flow_text
    ):
        text = m.group(1).strip()
        if text:
            points.append(
                KnowledgePoint(
                    topic=text[:24],
                    triggers=[],
                    key_points=[text[:200]],
                )
            )
    block_pattern = re.compile(
        r"\*\*([^*\n：:]{2,30})[：:]\*\*\s*([^\n]+)"
    )
    for m in block_pattern.finditer(call_flow_text):
        topic = m.group(1).strip()
        body = m.group(2).strip()
        if topic and body:
            points.append(
                KnowledgePoint(
                    topic=topic[:24],
                    triggers=[topic],
                    key_points=[body[:200]],
                )
            )
    return points


def parse_instruction(
    instruction_id: str,
    markdown: str,
    *,
    client: Optional[LLMClient] = None,
    model: Optional[str] = None,
    temperature: float = 0.1,
    mode: str = "llm",
) -> InstructionSpec:
    sections = _split_sections(markdown)
    if mode == "offline":
        return parse_instruction_offline(instruction_id, markdown)
    if client is None:
        client = build_default_client()
    messages = [
        ChatMessage(role="system", content=SYSTEM_PROMPT),
        ChatMessage(role="user", content=USER_TEMPLATE.format(markdown=markdown)),
    ]
    last_error: Optional[Exception] = None
    for attempt in range(3):
        try:
            result = client.chat(
                messages=messages,
                model=model,
                temperature=temperature,
                response_format={"type": "json_object"},
            )
            data = _extract_json(result.content)
            spec = _build_spec(instruction_id, markdown, sections, data)
            if not spec.constraints.hard.opening_keywords:
                spec.constraints.hard.opening_keywords = _offline_opening_keywords(
                    spec.opening_line_template
                )
            return spec
        except Exception as exc:
            last_error = exc
            logger.warning("Instruction parsing attempt %s failed: %s", attempt + 1, exc)
    logger.warning(
        "Falling back to offline parsing for instruction %s after %s", instruction_id, last_error
    )
    return parse_instruction_offline(instruction_id, markdown)


def parse_workbook(
    xlsx_path: str | Path,
    output_dir: str | Path = "data/instructions",
    *,
    client: Optional[LLMClient] = None,
    model: Optional[str] = None,
    mode: str = "llm",
) -> list[InstructionSpec]:
    rows = _load_rows(xlsx_path)
    if not rows:
        raise ValueError(f"No instruction rows found in {xlsx_path}")
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    specs: list[InstructionSpec] = []
    for instruction_id, markdown in rows:
        spec = parse_instruction(
            instruction_id, markdown, client=client, model=model, mode=mode
        )
        json_path = out_dir / f"{instruction_id}.json"
        md_path = out_dir / f"{instruction_id}.md"
        json_path.write_text(
            spec.model_dump_json(indent=2, by_alias=True), encoding="utf-8"
        )
        md_path.write_text(markdown, encoding="utf-8")
        specs.append(spec)
        logger.info(
            "Parsed instruction %s (variables=%s, nodes=%s, knowledge=%s)",
            instruction_id,
            spec.variables,
            len(spec.flow_nodes),
            len(spec.knowledge),
        )
    return specs


def load_specs(input_dir: str | Path = "data/instructions") -> list[InstructionSpec]:
    p = Path(input_dir)
    specs: list[InstructionSpec] = []
    for json_file in sorted(p.glob("*.json")):
        try:
            data = json.loads(json_file.read_text(encoding="utf-8"))
            specs.append(InstructionSpec.model_validate(data))
        except Exception:
            logger.exception("Failed to load instruction spec from %s", json_file)
    return specs
